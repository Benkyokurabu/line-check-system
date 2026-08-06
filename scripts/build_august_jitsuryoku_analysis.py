from __future__ import annotations

import html
import json
import statistics
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "●指導簿(2026年度) .xlsm"
OUTPUT_DIR = ROOT / "analysis_outputs"
OUTPUT = OUTPUT_DIR / "2026_08_jitsuryoku_analysis.html"
TARGET_YEAR = 2026
TARGET_MONTH = 8
TARGET_TEST = "前期実力テスト"
SUBJECTS = ["国語", "数学", "英語", "理科", "社会"]


def to_float(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def fmt(value):
    if value in (None, ""):
        return "-"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, float):
        return f"{value:.1f}"
    return str(value)


def date_parts(value):
    if hasattr(value, "year"):
        return value.year, value.month
    if value:
        try:
            parsed = datetime.fromisoformat(str(value)[:10])
            return parsed.year, parsed.month
        except ValueError:
            return None, None
    return None, None


def date_label(value):
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m")
    return str(value)[:7] if value else ""


def mean(values):
    nums = [v for v in values if v is not None]
    return round(sum(nums) / len(nums), 1) if nums else None


def median(values):
    nums = [v for v in values if v is not None]
    return round(statistics.median(nums), 1) if nums else None


def stdev(values):
    nums = [v for v in values if v is not None]
    return round(statistics.pstdev(nums), 1) if len(nums) >= 2 else None


def esc(value):
    return html.escape(str(value), quote=True)


def read_sheet_rows(wb, sheet_name, header_row, data_start):
    ws = wb[sheet_name]
    headers = [str(v) if v is not None else "" for v in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        yield dict(zip(headers, row))


def read_roster(wb):
    roster = {}
    for row in read_sheet_rows(wb, "入力_クラス一覧", 1, 2):
        sid = str(row.get("学籍番号") or "")
        if not sid:
            continue
        roster[sid] = {
            "student_id": sid,
            "grade": str(row.get("学年") or ""),
            "campus": str(row.get("所属") or ""),
            "name": str(row.get("本人氏名") or ""),
            "kana": str(row.get("ふりがな") or ""),
            "gender": str(row.get("性別") or ""),
            "school": str(row.get("学校") or ""),
            "teacher": str(row.get("担任") or ""),
            "classes": {subject: str(row.get(subject) or "") for subject in SUBJECTS},
            "notes": str(row.get("メモ") or ""),
        }
    return roster


def read_test_records(wb):
    records = []
    for row in read_sheet_rows(wb, "DB_塾内テスト", 3, 4):
        score = to_float(row.get("点数"))
        if score is None:
            continue
        year, month = date_parts(row.get("実施年月"))
        if not year:
            continue
        records.append(
            {
                "student_id": str(row.get("学籍番号") or ""),
                "name": str(row.get("氏名") or ""),
                "test": str(row.get("テスト名") or ""),
                "date": date_label(row.get("実施年月")),
                "year": year,
                "month": month,
                "campus": str(row.get("校舎") or ""),
                "grade": str(row.get("学年") or ""),
                "subject": str(row.get("科目") or ""),
                "klass": str(row.get("クラス") or ""),
                "lesson": str(row.get("授業名") or ""),
                "score": score,
                "deviation": to_float(row.get("塾内偏差値")),
                "rank": to_float(row.get("順位")),
                "count": to_float(row.get("人数")),
            }
        )
    return records


def read_hokushin(wb):
    rows_by_student = defaultdict(list)
    for row in read_sheet_rows(wb, "DB_外部模試", 4, 5):
        sid = str(row.get("学籍番号") or "")
        test = str(row.get("テスト名") or "")
        if not sid or "北辰" not in test:
            continue
        y, m = date_parts(row.get("実施日"))
        record = {
            "student_id": sid,
            "name": str(row.get("氏名") or ""),
            "test": test,
            "date": date_label(row.get("実施日")),
            "year": y or 0,
            "month": m or 0,
            "three_score": to_float(row.get("三科点")),
            "three_dev": to_float(row.get("三科偏差")),
            "five_score": to_float(row.get("五科点")),
            "five_dev": to_float(row.get("五科偏差")),
            "subjects": {},
        }
        for subject in SUBJECTS:
            record["subjects"][subject] = {
                "score": to_float(row.get(f"{subject}点")),
                "dev": to_float(row.get(f"{subject}偏差")),
            }
        rows_by_student[sid].append(record)
    latest = {}
    for sid, items in rows_by_student.items():
        latest[sid] = sorted(items, key=lambda r: (r["year"], r["month"], r["test"]), reverse=True)[0]
    return latest


def group_stats(records, keys):
    groups = defaultdict(list)
    for record in records:
        groups[tuple(record[k] for k in keys)].append(record)
    rows = []
    for key, items in groups.items():
        scores = [r["score"] for r in items]
        rows.append({**{keys[i]: key[i] for i in range(len(keys))}, "n": len(items), "avg": mean(scores), "median": median(scores), "stdev": stdev(scores), "max": max(scores), "min": min(scores), "under40": sum(1 for s in scores if s < 40), "over80": sum(1 for s in scores if s >= 80)})
    return sorted(rows, key=lambda r: tuple(str(r[k]) for k in keys))


def find_recent_tests(records):
    target_order = (TARGET_YEAR, TARGET_MONTH)
    by_grade_subject = defaultdict(dict)
    for r in records:
        if (r["year"], r["month"]) >= target_order or r["test"] == TARGET_TEST:
            continue
        key = (r["grade"], r["subject"])
        by_grade_subject[key][(r["year"], r["month"], r["test"], r["date"])] = True
    result = {}
    for key, tests in by_grade_subject.items():
        result[key] = sorted(tests.keys(), reverse=True)[:2]
    return result


def build_students(records, roster, hokushin):
    current = [r for r in records if r["year"] == TARGET_YEAR and r["month"] == TARGET_MONTH and r["test"] == TARGET_TEST]
    recent_defs = find_recent_tests(records)
    by_student = defaultdict(list)
    past_lookup = defaultdict(dict)
    for r in records:
        sid = r["student_id"]
        if r in current:
            by_student[sid].append(r)
        test_key = (r["year"], r["month"], r["test"], r["date"])
        past_lookup[(sid, r["grade"], r["subject"])][test_key] = r

    students = []
    for sid, items in by_student.items():
        base = roster.get(sid, {})
        name = base.get("name") or items[0]["name"]
        grade = base.get("grade") or items[0]["grade"]
        campus = base.get("campus") or ("本" if items[0]["campus"] == "本校" else "南" if items[0]["campus"] == "南教室" else items[0]["campus"])
        current_by_subject = {r["subject"]: r for r in items}
        current_scores = [r["score"] for r in current_by_subject.values()]
        row = {
            "student_id": sid,
            "name": name,
            "kana": base.get("kana", ""),
            "grade": grade,
            "campus": campus,
            "school": base.get("school", ""),
            "teacher": base.get("teacher", ""),
            "gender": base.get("gender", ""),
            "current_classes": base.get("classes", {}),
            "subjects": {},
            "current_total": round(sum(current_scores), 1),
            "current_avg": round(sum(current_scores) / len(current_scores), 1),
            "tested": len(current_scores),
            "hokushin": hokushin.get(sid),
            "flags": [],
            "suggestion_score": 0,
        }
        for subject in SUBJECTS:
            current_record = current_by_subject.get(subject)
            subject_recent = []
            for test_key in recent_defs.get((grade, subject), []):
                past = past_lookup.get((sid, grade, subject), {}).get(test_key)
                subject_recent.append({"label": f"{test_key[3]} {test_key[2]}", "score": past["score"] if past else None, "dev": past["deviation"] if past else None})
            hs = hokushin.get(sid, {}).get("subjects", {}).get(subject, {}) if hokushin.get(sid) else {}
            row["subjects"][subject] = {
                "class": base.get("classes", {}).get(subject, ""),
                "current": current_record["score"] if current_record else None,
                "current_dev": current_record["deviation"] if current_record else None,
                "current_rank": current_record["rank"] if current_record else None,
                "recent": subject_recent,
                "hokushin_score": hs.get("score"),
                "hokushin_dev": hs.get("dev"),
            }
        low_subjects = [s for s, v in row["subjects"].items() if v["current"] is not None and v["current"] < 40]
        high_subjects = [s for s, v in row["subjects"].items() if v["current"] is not None and v["current"] >= 80]
        if row["current_avg"] < 50:
            row["flags"].append("平均50未満")
        if low_subjects:
            row["flags"].append("40未満: " + "/".join(low_subjects))
        if high_subjects:
            row["flags"].append("80以上: " + "/".join(high_subjects))
        if row["hokushin"] and row["hokushin"].get("three_dev") is not None:
            row["suggestion_score"] = round(row["current_avg"] * 0.65 + row["hokushin"]["three_dev"] * 1.2, 1)
        else:
            row["suggestion_score"] = row["current_avg"]
        students.append(row)
    return sorted(students, key=lambda r: (r["grade"], r["campus"], -r["suggestion_score"], r["kana"], r["name"]))


def render_table(rows, columns):
    thead = "".join(f"<th>{esc(label)}</th>" for _, label in columns)
    body = []
    for row in rows:
        body.append("<tr>" + "".join(f"<td>{esc(fmt(row.get(key, '')))}</td>" for key, _ in columns) + "</tr>")
    return f"<table><thead><tr>{thead}</tr></thead><tbody>{''.join(body)}</tbody></table>"


def class_select(student, subject):
    current = student["subjects"].get(subject, {}).get("class") or ""
    options = ["", "Ｓ", "Ａ", "Ｂ", "Ｃ", "X", "個", "非受講"]
    option_html = []
    for option in options:
        selected = " selected" if option == current else ""
        label = option or "未定"
        option_html.append(f'<option value="{esc(option)}"{selected}>{esc(label)}</option>')
    return f'<select class="class-change" data-subject="{esc(subject)}" data-current="{esc(current)}">{"".join(option_html)}</select>'


def placement_cells(student):
    cells = []
    for subject in ["国語", "数学", "英語"]:
        s = student["subjects"].get(subject, {})
        recent = s.get("recent", [])
        r1 = recent[0] if len(recent) > 0 else {}
        r2 = recent[1] if len(recent) > 1 else {}
        cells.extend(
            [
                {"value": fmt(s.get("class"))},
                {"value": class_select(student, subject), "html": True},
                {"value": fmt(s.get("current")), "sort": s.get("current")},
                {"value": fmt(s.get("current_dev")), "sort": s.get("current_dev")},
                {"value": fmt(r1.get("score")), "sort": r1.get("score")},
                {"value": fmt(r2.get("score")), "sort": r2.get("score")},
                {"value": fmt(s.get("hokushin_dev")), "sort": s.get("hokushin_dev")},
            ]
        )
    hs = student.get("hokushin") or {}
    cells.extend(
        [
            {"value": fmt(hs.get("test"))},
            {"value": fmt(hs.get("three_score")), "sort": hs.get("three_score")},
            {"value": fmt(hs.get("three_dev")), "sort": hs.get("three_dev")},
            {"value": fmt(hs.get("five_score")), "sort": hs.get("five_score")},
            {"value": fmt(hs.get("five_dev")), "sort": hs.get("five_dev")},
        ]
    )
    return cells


def make_td(cell):
    if isinstance(cell, dict):
        sort = cell.get("sort")
        sort_attr = "" if sort in (None, "") else f' data-sort="{esc(sort)}"'
        value = cell.get("value", "")
        if cell.get("html"):
            return f"<td{sort_attr}>{value}</td>"
        return f"<td{sort_attr}>{esc(value)}</td>"
    return f"<td>{esc(cell)}</td>"


def make_placement_table(students, table_id="placementTable"):
    fixed = ["学年", "校舎", "学籍番号", "氏名", "学校", "担任", "現平均", "判定用", "注意"]
    subject_cols = []
    for subject in ["国語", "数学", "英語"]:
        subject_cols += [f"{subject}現クラス", f"{subject}変更案", f"{subject}実力テスト", f"{subject}実力偏", f"{subject}単元テスト②", f"{subject}単元テスト①", f"{subject}北辰偏"]
    tail = ["北辰回", "北辰3科", "北辰3科偏", "北辰5科", "北辰5科偏"]
    header = "".join(f"<th>{esc(x)}</th>" for x in fixed + subject_cols + tail)
    rows = []
    for st in students:
        base = [
            {"value": st["grade"]},
            {"value": st["campus"]},
            {"value": st["student_id"]},
            {"value": st["name"]},
            {"value": st["school"]},
            {"value": st["teacher"]},
            {"value": fmt(st["current_avg"]), "sort": st["current_avg"]},
            {"value": fmt(st["suggestion_score"]), "sort": st["suggestion_score"]},
            {"value": " / ".join(st["flags"])},
        ]
        search = st["student_id"] + " " + st["name"] + " " + st["kana"] + " " + st["school"]
        cells = "".join(make_td(cell) for cell in base + placement_cells(st))
        rows.append(f'<tr data-grade="{esc(st["grade"])}" data-campus="{esc(st["campus"])}" data-search="{esc(search)}">{cells}</tr>')
    return f'<table id="{esc(table_id)}" class="placement-table sortable"><thead><tr>{header}</tr></thead><tbody>{"".join(rows)}</tbody></table>'

def distribution(records):
    bands = [(80, "80点以上"), (60, "60-79点"), (40, "40-59点"), (0, "39点以下")]
    groups = defaultdict(lambda: defaultdict(int))
    for r in records:
        for floor, label in bands:
            if r["score"] >= floor:
                groups[(r["grade"], r["campus"], r["subject"])][label] += 1
                break
    rows = []
    for key, counts in groups.items():
        total = sum(counts.values())
        rows.append({"grade": key[0], "campus": key[1], "subject": key[2], "total": total, **{label: counts.get(label, 0) for _, label in bands}})
    return sorted(rows, key=lambda r: (r["grade"], r["campus"], r["subject"]))


def main():
    wb = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True, keep_vba=True)
    roster = read_roster(wb)
    all_records = read_test_records(wb)
    current_records = [r for r in all_records if r["year"] == TARGET_YEAR and r["month"] == TARGET_MONTH and r["test"] == TARGET_TEST]
    hokushin = read_hokushin(wb)
    students = build_students(all_records, roster, hokushin)
    concern = [s for s in students if s["flags"]]
    subject_stats = group_stats(current_records, ["grade", "campus", "subject"])
    class_stats = group_stats(current_records, ["grade", "campus", "subject", "klass"])
    dist = distribution(current_records)

    payload = {"students": students, "subjectStats": subject_stats, "classStats": class_stats, "distribution": dist}
    metrics = [("延べ件数", len(current_records), "今回テスト"), ("受験者", len(students), "1科目以上"), ("全体平均", fmt(mean([r["score"] for r in current_records])), "全学年"), ("北辰あり", sum(1 for s in students if s.get("hokushin")), "最新回を表示"), ("要確認", len(concern), "平均50未満等")]
    metrics_html = "".join(f"<section class=\"metric\"><span>{esc(a)}</span><strong>{esc(b)}</strong><em>{esc(c)}</em></section>" for a, b, c in metrics)
    placement_html = make_placement_table(students)
    subject_html = render_table(subject_stats, [("grade", "学年"), ("campus", "校舎"), ("subject", "科目"), ("n", "人数"), ("avg", "平均"), ("median", "中央値"), ("stdev", "標準偏差"), ("max", "最高"), ("min", "最低"), ("under40", "40未満"), ("over80", "80以上")])
    class_html = render_table(class_stats, [("grade", "学年"), ("campus", "校舎"), ("subject", "科目"), ("klass", "現クラス"), ("n", "人数"), ("avg", "平均"), ("median", "中央値"), ("stdev", "標準偏差"), ("max", "最高"), ("min", "最低"), ("under40", "40未満"), ("over80", "80以上")])
    concern_html = make_placement_table(concern, "concernPlacementTable")
    dist_rows = []
    for r in dist:
        total = r["total"] or 1
        bar = "".join(f"<i class=\"b{i}\" style=\"width:{r[label] / total * 100:.2f}%\"></i>" for i, label in enumerate(["80点以上", "60-79点", "40-59点", "39点以下"], 1))
        dist_rows.append(f"<tr><td>{esc(r['grade'])}</td><td>{esc(r['campus'])}</td><td>{esc(r['subject'])}</td><td>{r['total']}</td><td><div class=\"bar\">{bar}</div></td><td>{r['80点以上']}</td><td>{r['60-79点']}</td><td>{r['40-59点']}</td><td>{r['39点以下']}</td></tr>")

    html_text = f'''<!doctype html>
<html lang="ja">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>2026年8月 前期実力テスト クラス替え資料</title>
<style>
:root {{ --ink:#172033; --muted:#637083; --line:#d7dce5; --bg:#f5f7fa; --panel:#fff; --head:#eef3f8; --blue:#2f6fdf; --green:#138a57; --yellow:#bc7a00; --red:#c43d3d; --header-height:58px; --top-controls-height:170px; --panel-toolbar-height:90px; --panel-toolbar-top:calc(var(--header-height) + var(--top-controls-height)); --table-head-top:calc(var(--header-height) + var(--top-controls-height) + var(--panel-toolbar-height)); }}
* {{ box-sizing:border-box; }} body {{ margin:0; font-family:"Yu Gothic","Meiryo",system-ui,sans-serif; color:var(--ink); background:var(--bg); font-size:13px; }}
header {{ position:sticky; top:0; z-index:40; background:var(--panel); border-bottom:1px solid var(--line); padding:14px 18px 10px; }}
h1 {{ margin:0; font-size:20px; letter-spacing:0; }} .meta {{ color:var(--muted); margin-top:3px; }} main {{ max-width:1600px; margin:0 auto; padding:16px 18px 44px; }} .top-controls {{ position:sticky; top:var(--header-height); z-index:35; background:var(--bg); padding-bottom:1px; }}
.metrics {{ display:grid; grid-template-columns:repeat(5,minmax(120px,1fr)); gap:10px; margin-bottom:12px; }} .metric {{ background:var(--panel); border:1px solid var(--line); border-radius:8px; padding:10px 12px; }} .metric span,.metric em {{ display:block; color:var(--muted); font-style:normal; }} .metric strong {{ display:block; font-size:24px; margin:1px 0; }}
.filters {{ display:flex; flex-wrap:wrap; gap:10px; align-items:end; background:var(--panel); border:1px solid var(--line); border-radius:8px; padding:10px; margin-bottom:12px; }} label {{ display:grid; gap:3px; color:var(--muted); }} select,input {{ height:32px; border:1px solid var(--line); border-radius:6px; padding:0 8px; background:#fff; }} input {{ min-width:260px; }} .class-change {{ min-width:64px; height:28px; padding:0 5px; }}
.tabs {{ display:flex; gap:6px; border-bottom:1px solid var(--line); margin-top:12px; }} .tab {{ border:1px solid var(--line); border-bottom:0; background:#e9eef5; padding:8px 12px; border-radius:6px 6px 0 0; cursor:pointer; }} .tab.active {{ background:var(--panel); font-weight:700; }}
.panel {{ display:none; background:var(--panel); border:1px solid var(--line); border-top:0; padding:12px; overflow:auto; }} .panel.active {{ display:block; }} .panel-toolbar {{ position:sticky; top:var(--panel-toolbar-top); z-index:30; background:var(--panel); border-bottom:1px solid var(--line); padding:0 0 10px; margin:0 0 10px; }} h2 {{ font-size:16px; margin:0 0 8px; }} .note {{ color:var(--muted); margin:0 0 10px; }}
table {{ border-collapse:separate; border-spacing:0; width:100%; min-width:1200px; }} th,td {{ border-right:1px solid #e3e8ef; border-bottom:1px solid #e3e8ef; padding:6px 7px; white-space:nowrap; text-align:left; }} th {{ position:sticky; top:var(--table-head-top); z-index:5; background:var(--head); font-weight:700; color:#2d3748; }} td:nth-child(n+7) {{ text-align:right; }} tbody tr:hover {{ background:#f8fbff; }}
.placement-table th:nth-child(-n+4),.placement-table td:nth-child(-n+4) {{ position:sticky; z-index:4; }} .placement-table th:nth-child(1),.placement-table td:nth-child(1) {{ left:0; min-width:52px; width:52px; }} .placement-table th:nth-child(2),.placement-table td:nth-child(2) {{ left:52px; min-width:48px; width:48px; }} .placement-table th:nth-child(3),.placement-table td:nth-child(3) {{ left:100px; min-width:86px; width:86px; }} .placement-table th:nth-child(4),.placement-table td:nth-child(4) {{ left:186px; min-width:142px; width:142px; }} .placement-table td:nth-child(-n+4) {{ background:#fff; }} .placement-table tbody tr:hover td:nth-child(-n+4) {{ background:#f8fbff; }} .placement-table th:nth-child(-n+4) {{ background:var(--head); z-index:8; }}
.bar {{ display:flex; width:260px; height:18px; border-radius:4px; overflow:hidden; background:#e5e7eb; }} .b1 {{ background:var(--green); }} .b2 {{ background:var(--blue); }} .b3 {{ background:var(--yellow); }} .b4 {{ background:var(--red); }} .hidden-row {{ display:none; }} th.sortable-head {{ cursor:pointer; user-select:none; }} th.sortable-head::after {{ content:" ⇅"; color:var(--muted); font-weight:400; }} th.sort-asc::after {{ content:" ↑"; color:var(--blue); }} th.sort-desc::after {{ content:" ↓"; color:var(--blue); }} .class-counts {{ display:grid; grid-template-columns:repeat(auto-fit,minmax(220px,1fr)); gap:10px; margin:10px 0 12px; }} .count-card {{ border:1px solid var(--line); border-radius:8px; padding:9px 10px; background:#fbfcfe; }} .count-card h3 {{ margin:0 0 6px; font-size:13px; }} .chips {{ display:flex; flex-wrap:wrap; gap:5px; }} .chip {{ border:1px solid var(--line); border-radius:999px; padding:2px 7px; background:#fff; }} .changed {{ background:#fff7ed; }}
@media print {{ header,.filters,.tabs {{ display:none; }} main {{ padding:0; }} .panel {{ display:block; border:0; }} .panel:not(.active) {{ display:none; }} th {{ position:static; }} }}
@media (max-width:760px) {{ header {{ position:static; }} .top-controls,.panel-toolbar {{ position:static; }} main {{ padding:12px; }} .metrics {{ grid-template-columns:repeat(2,minmax(120px,1fr)); }} input,select {{ width:100%; }} label {{ flex:1 1 140px; }} th {{ top:0; }} }}
</style>
</head>
<body>
<header><h1>2026年8月 前期実力テスト クラス替え資料</h1><div class="meta">参照元: {esc(SOURCE.name)} / 作成: {datetime.now().strftime('%Y-%m-%d %H:%M')} / 直近2回は今回より前の同学年・同科目から自動取得</div></header>
<main>
<div class="top-controls">
<div class="metrics">{metrics_html}</div>
<section class="filters"><label>学年<select id="gradeFilter"><option value="">全て</option></select></label><label>校舎<select id="campusFilter"><option value="">全て</option></select></label><label>検索<input id="searchFilter" type="search" placeholder="氏名・ふりがな・学籍番号・学校"></label></section>
<div class="tabs"><button class="tab active" data-panel="placement">クラス替え資料</button><button class="tab" data-panel="concern">要確認</button><button class="tab" data-panel="subject">科目別</button><button class="tab" data-panel="class">現クラス別</button><button class="tab" data-panel="dist">分布</button></div>
</div>
<section id="placement" class="panel active"><div class="panel-toolbar"><h2>クラス替え資料</h2><p class="note">変更案はこの画面上だけの作業用です。人数は校舎別に集計され、変更すると即時更新されます。列見出しをクリックすると各テスト・偏差値でソートできます。</p><div id="classCounts" class="class-counts"></div></div>{placement_html}</section>
<section id="concern" class="panel"><div class="panel-toolbar"><h2>要確認</h2><p class="note">平均50未満、40点未満科目、または80点以上科目がある生徒を抽出しています。</p></div>{concern_html}</section>
<section id="subject" class="panel"><div class="panel-toolbar"><h2>学年・校舎・科目別</h2></div>{subject_html}</section>
<section id="class" class="panel"><div class="panel-toolbar"><h2>現クラス別</h2></div>{class_html}</section>
<section id="dist" class="panel"><div class="panel-toolbar"><h2>点数分布</h2></div><table><thead><tr><th>学年</th><th>校舎</th><th>科目</th><th>人数</th><th>分布</th><th>80以上</th><th>60-79</th><th>40-59</th><th>39以下</th></tr></thead><tbody>{''.join(dist_rows)}</tbody></table></section>
</main>
<script id="payload" type="application/json">{json.dumps(payload, ensure_ascii=False)}</script>
<script>
const data = JSON.parse(document.getElementById('payload').textContent);
const gradeFilter = document.getElementById('gradeFilter');
const campusFilter = document.getElementById('campusFilter');
const searchFilter = document.getElementById('searchFilter');
const classCounts = document.getElementById('classCounts');
const classOrder = ['', 'Ｓ', 'Ａ', 'Ｂ', 'Ｃ', 'X', '個', '非受講'];
function updateStickyOffsets() {{
  const header = document.querySelector('header');
  const topControls = document.querySelector('.top-controls');
  const activeToolbar = document.querySelector('.panel.active .panel-toolbar');
  const headerHeight = header ? Math.ceil(header.getBoundingClientRect().height) : 0;
  const topControlsHeight = topControls ? Math.ceil(topControls.getBoundingClientRect().height) : 0;
  const toolbarHeight = activeToolbar ? Math.ceil(activeToolbar.getBoundingClientRect().height) : 0;
  document.documentElement.style.setProperty('--header-height', `${{headerHeight}}px`);
  document.documentElement.style.setProperty('--top-controls-height', `${{topControlsHeight}}px`);
  document.documentElement.style.setProperty('--panel-toolbar-height', `${{toolbarHeight}}px`);
}}
function addOptions(select, values) {{
  [...new Set(values.filter(Boolean))].sort().forEach(v => {{
    const o = document.createElement('option');
    o.value = v;
    o.textContent = v;
    select.appendChild(o);
  }});
}}
addOptions(gradeFilter, data.students.map(s => s.grade));
addOptions(campusFilter, data.students.map(s => s.campus));
document.querySelectorAll('.tab').forEach(button => button.addEventListener('click', () => {{
  document.querySelectorAll('.tab').forEach(x => x.classList.remove('active'));
  document.querySelectorAll('.panel').forEach(x => x.classList.remove('active'));
  button.classList.add('active');
  document.getElementById(button.dataset.panel).classList.add('active');
  updateStickyOffsets();
  applyFilters();
}}));
function visiblePlacementRows() {{
  return [...document.querySelectorAll('#placementTable tbody tr')].filter(row => !row.classList.contains('hidden-row'));
}}
function updateClassCounts() {{
  if (!classCounts) return;
  const counts = {{}};
  visiblePlacementRows().forEach(row => {{
    const campus = row.dataset.campus || '校舎未設定';
    row.querySelectorAll('.class-change').forEach(select => {{
      const subject = select.dataset.subject;
      const value = select.value || '未定';
      counts[campus] ??= {{}};
      counts[campus][subject] ??= {{}};
      counts[campus][subject][value] = (counts[campus][subject][value] || 0) + 1;
      select.closest('td').classList.toggle('changed', select.value !== select.dataset.current);
    }});
  }});
  const campuses = Object.keys(counts).sort((a, b) => a.localeCompare(b, 'ja', {{ numeric: true }}));
  classCounts.innerHTML = campuses.flatMap(campus => ['国語','数学','英語'].map(subject => {{
    const subjectCounts = counts[campus]?.[subject] || {{}};
    const chips = [...new Set([...classOrder.map(x => x || '未定'), ...Object.keys(subjectCounts)])]
      .filter(k => subjectCounts[k])
      .map(k => `<span class="chip">${{k}}: ${{subjectCounts[k]}}</span>`)
      .join('');
    return `<section class="count-card"><h3>${{campus}} ${{subject}} 変更案人数</h3><div class="chips">${{chips || '<span class="chip">対象なし</span>'}}</div></section>`;
  }})).join('') || '<section class="count-card"><h3>変更案人数</h3><div class="chips"><span class="chip">対象なし</span></div></section>';
  updateStickyOffsets();
}}
function applyFilters() {{
  const g = gradeFilter.value;
  const c = campusFilter.value;
  const q = searchFilter.value.trim();
  document.querySelectorAll('tbody tr[data-grade]').forEach(row => {{
    const ok = (!g || row.dataset.grade === g) && (!c || row.dataset.campus === c) && (!q || row.dataset.search.includes(q));
    row.classList.toggle('hidden-row', !ok);
  }});
  updateClassCounts();
}}
function cellSortValue(cell) {{
  const explicit = cell.dataset.sort;
  if (explicit !== undefined && explicit !== '') return explicit;
  const select = cell.querySelector('select');
  return select ? select.value : cell.textContent.trim();
}}
function compareValues(a, b) {{
  const av = cellSortValue(a);
  const bv = cellSortValue(b);
  const an = Number(String(av).replace(/,/g, ''));
  const bn = Number(String(bv).replace(/,/g, ''));
  if (!Number.isNaN(an) && !Number.isNaN(bn) && String(av).trim() !== '' && String(bv).trim() !== '') return an - bn;
  return String(av).localeCompare(String(bv), 'ja', {{ numeric: true }});
}}
function makeSortable(table) {{
  const headers = [...table.querySelectorAll('thead th')];
  headers.forEach((th, index) => {{
    th.classList.add('sortable-head');
    th.addEventListener('click', () => {{
      const direction = th.classList.contains('sort-asc') ? -1 : 1;
      headers.forEach(h => h.classList.remove('sort-asc', 'sort-desc'));
      th.classList.add(direction === 1 ? 'sort-asc' : 'sort-desc');
      const tbody = table.querySelector('tbody');
      const rows = [...tbody.querySelectorAll('tr')];
      rows.sort((ra, rb) => compareValues(ra.children[index], rb.children[index]) * direction);
      rows.forEach(row => tbody.appendChild(row));
      applyFilters();
    }});
  }});
}}
document.querySelectorAll('table').forEach(makeSortable);
document.querySelectorAll('.class-change').forEach(select => select.addEventListener('change', updateClassCounts));
[gradeFilter, campusFilter, searchFilter].forEach(x => x.addEventListener('input', () => {{ updateStickyOffsets(); applyFilters(); }}));
window.addEventListener('resize', () => {{ updateStickyOffsets(); applyFilters(); }});
updateStickyOffsets();
applyFilters();
</script>
</body>
</html>'''
    OUTPUT_DIR.mkdir(exist_ok=True)
    OUTPUT.write_text(html_text, encoding="utf-8")
    print(OUTPUT)
    print(f"records={len(current_records)} students={len(students)} hokushin={sum(1 for s in students if s.get('hokushin'))} concerns={len(concern)}")


if __name__ == "__main__":
    main()

